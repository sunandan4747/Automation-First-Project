import openpyxl


class HomePageData:     ## class name has to be pascle type

    test_HomePage_data = [{"firstname":"Sagar","lastname":"Sarade","gender":"Male"}, {"firstname":"Yusuf", "lastname":"Tamboli", "gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):        # method name has to be camel case
        Dict = {}
        book = openpyxl.load_workbook("../TestData/UserData.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

